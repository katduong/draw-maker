from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font


def parseFile():
    wb = load_workbook('players.xlsx')
    main = wb['Players']

    flights = {
        'AMD': 0, 'BMD': 0,'CMD': 0,'DMD': 0,'AWD': 0,
        'BWD': 0,'CWD': 0,'DWD': 0,'AMX': 0,'BMX': 0,
        'CMX': 0,'DMX': 0,'AMS': 0,'BMS': 0,'CMS': 0,
        'DMS': 0,'AWS': 0,'BWS': 0,'CWS': 0,'DWS': 0,
        'AXD': 0,'BXD': 0,'CXD': 0,'DXD': 0
    }

    header = list(main.rows)[3]
    for cell in header:
        cell.font = Font(bold=True)
    # loop through every row and create a new sheet for each event
    # player will be appended to the event's sheet
    for row in main.rows:
        print(row[4].value)
        events = row[4].value
        # skip first couple lines
        if (events != None):
            eventList = events.split(", ")
            print(eventList)
            for event in eventList:
                if event in flights:
                    if flights[event] == 0:
                        flights[event] += 1
                        newSheet = wb.create_sheet(event)
                        newSheet.append(cell.value for cell in header)
                        newSheet.append(cell.value for cell in row)
                    else:
                        wb[event].append(cell.value for cell in row)
    # sort sheets in alphabetical order
    wb._sheets.sort(key=lambda ws:ws.title)
    wb.save('playerParse.xlsx')

def reformatParse():
    f = open("errors.txt", "w+")
    wb = load_workbook("playerParse.xlsx")
    for sheet in wb:
        f.write("******************************* " + sheet.title + " **********************************\n")
        if sheet.title != "Players":
            # insert column for event letters (ex: AB, BC,... )
            sheet.insert_cols(5, amount=1)
            # if sheet is singles, don't need to worry about partners
            # just need to add reformat columns
            # lastname, firstname, gender, event letter(s), club


            # For each row in the sheet, create a player events dictionary
            # and store the letters of the all events the person is playing
            for row in sheet:
                playerEvents = {
                    'MS': "",

                    'WS': "",
                    'MX': "",
                    'MD': "",
                    'WD': "",
                    'XD': ""
                }
                events = row[5].value.split(", ")
                # exclude first row
                if events != ['Events']:
                    for event in events:
                        # store the last two letters of event in variable string (MD, MX, WD, ...)
                        string = event[-2:]
                        # add flight letters to each event in playerEvents dictionary using string as the key
                        playerEvents[string] += event[0]
                    # store the flights the player is playing for this event in the new events column
                    row[4].value = playerEvents[sheet.title[1:]]
                # make the title for the events column
                else:
                    row[4].value = "Events"
                print(sheet.title)
            # doubles event
            # need to create a column for doubles partners
            # lastname, firstname, gender, event letter(s), doubles partner, club
            if 'MD' in sheet.title or 'WD' in sheet.title or 'MX' in sheet.title or 'XD' in sheet.title:
                # insert a column for the partner name
                sheet.insert_cols(6, amount=1)
                # insert a column for the partner club
                sheet.insert_cols(7, amount=1)
                players = {}
                rowCount = 1
                for row in sheet:
                    # skip first row for title
                    if rowCount == 1:
                        rowCount += 1
                        continue
                    # grab entry info for that player
                    entryInfo = row[8].value
                    # store first name and last name in currentPlayer
                    currentPlayer = row[1].value + " " + row[0].value
                    print("*"*20)
                    print(currentPlayer)
                    # print("entry info:", entryInfo)
                    # entry info is not empty, split into an array
                    if entryInfo != None:
                        # gets entry info and splits into list per each event
                        entryInfo = entryInfo.split("\n")
                        print("entry info:", entryInfo)
                        # remove empty string from the last index of list
                        if entryInfo[-1] == "":
                            entryInfo.pop()
                        print("entry info after pop()",entryInfo)
                        # loop through each event in the player's entry entryInfo
                        # check if the flight is the current event flight
                        # store their partner's name in player dictionary
                        for event in entryInfo:
                            if sheet.title in event:
                                print("event:",event)
                                # partnerName contains first and last name
                                # omit first four characters and last 3 characters
                                # partnerName = event[4:-3]#.split(" ")
                                partnerName = event[4:].title().split(" (")[0]
                                partnerName = partnerName.split(" ")
                                print("partner name:",partnerName)
                                # stores each player as a key in the dictionary
                                # values will be a list of lists
                                # [[partner fn, ln], [club, rowIndex of partner]]
                                # players[currentPlayer] = []
                                players[currentPlayer] = []
                                # players[currentPlayer].append([])
                                # for name in partnerName:
                                    # players[currentPlayer][0].append(name)
                                players[currentPlayer].append([])
                                for name in partnerName:
                                    players[currentPlayer][0].append(name)
                                club = row[3].value
                                players[currentPlayer].append([])
                                players[currentPlayer][1].append(club)
                                # keep track of where the
                                players[currentPlayer][1].append(rowCount)
                    rowCount+=1
                print(players)
                # iterate through each row again and check if player is listed in the dictionary as another player's partners
                # if a match is found, append to the first occurence of pair's dictionary, the partner's club name
                rowCount = 1
                partnerDict = {}
                for row in sheet:
                    if rowCount == 1:
                        rowCount += 1
                        continue
                    playerName = row[1].value + " " + row[0].value
                    if playerName not in players:
                        f.write(playerName + " not found in player dictionary\n")
                        print(playerName + " not found in player dictionary\n")
                        rowCount += 1
                        continue
                    # if player's partner in dictionary doesn't have a first name or last name, skip
                    if players[playerName][0][0] == "":
                        f.write(playerName + " left their partner blank\n")
                        rowCount += 1
                        continue
                    #  check if player's partner only has a first name listed
                    elif players[playerName][0][-1] == players[playerName][0][0]:
                        partnerfn = players[playerName][0][0]
                        f.write("Check " + playerName + "'s partner " + partnerfn + "\n")
                        # partnerFullName = partnerfn
                    #  otherwise, we know their partner has a first and last name
                    else:
                        partnerfn = players[playerName][0][0]
                        partnerln = players[playerName][0][-1]
                        partnerFullName = partnerfn + " " + partnerln

                    # check if partner full name matches with a key in the dictionary
                    if partnerFullName in players:
                        playerFnInDict = players[partnerFullName][0][0]
                        playerLnInDict = players[partnerFullName][0][-1]
                        # if playerFnInDict == row[1].value:
                        # store partner name in excel sheet
                        row[5].value = partnerFullName
                        if players[partnerFullName][1][0] != None:
                            # store partner's club name in excel sheet
                            row[6].value = players[partnerFullName][1][0]
                        # store row number of partner and row number of player in dictionary so that we can delete the row later
                        if not rowCount in partnerDict.values():
                            partnerDict[rowCount] = players[partnerFullName][1][-1]
                    # partner name has fn and ln but spelled incorrectly
                    else:
                        f.write(playerName + " spelled their partner's name incorrectly: " + partnerFullName + "\n")
                    rowCount += 1
                #     rowCount += 1
                print(partnerDict)

                # loop through items in dictionary and delete corresponding row
                # also delete other pair in dictionary
                for playerRow, partnerRow in sorted(partnerDict.items(), reverse=True):
                    print("deleting row", playerRow)
                    sheet.delete_rows(playerRow)
    wb.save("playerList.xlsx")
    f.close()


def main():
    # parseFile()
    reformatParse()


if __name__ == '__main__':
    main()
