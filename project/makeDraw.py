from openpyxl import load_workbook, Workbook
from operator import itemgetter
import random


def makeDraw():
    wb = load_workbook("masterlist.xlsx")
    sheet = wb['DMS']

    rowCount = 1
    lfplayers = []
    midplayers = []
    hfplayers = []
    flightDict = {
        'A': ['A','AB'],
        'B': ['AB','B','BC'],
        'C': ['BC','C','CD'],
        'D': ['CD','D']
    }
    letter = sheet.title[0]
    flights = flightDict[letter]
    print(flights)

    for row in sheet:
        if row[0].value == 'Last Name':
            continue
        if row[3].value == flights[0]:
            ln = row[0].value
            fn = row[1].value
            club = row[4].value
            flight = row[3].value
            if club == None:
                club='Z'
            hfplayers.append([fn,ln,club,flight])
        if row[3].value == flights[1]:
            ln = row[0].value
            fn = row[1].value
            club = row[4].value
            flight = row[3].value

            if club == None:
                club='Z'
            midplayers.append([fn,ln,club,flight])
        if letter == 'B' or letter == 'C':
            if row[3].value == flights[2]:
                ln = row[0].value
                fn = row[1].value
                club = row[4].value
                flight = row[3].value
                if club == None:
                    club='Z'
                lfplayers.append([fn,ln,club,flight])
    midplayers = sorted(midplayers, key=itemgetter(2))
    lfplayers = sorted(lfplayers, key=itemgetter(2))
    hfplayers = sorted(hfplayers, key=itemgetter(2))

    players = hfplayers + midplayers + lfplayers
    print("*"*20)
    print("all players:", players)

    numRows = len(hfplayers) + len(midplayers) + len(lfplayers)
    print(numRows)
    # if numPlayers > 32 and numPlayers <  64:
    firstquadrant = [None]*8
    secondquadrant = [None]*8
    thirdquadrant = [None]*8
    fourthquadrant = [None]*8
    # number of matches that can be filled on the second round minus pullouts
    if numRows >= 64:
        smallerBracket = 64
    elif numRows >= 32 and numRows < 64:
        smallerBracket = 32
    elif numRows >= 16 and numRows < 32:
        smallerBracket = 16
    elif numRows >= 8 and numRows < 16:
        smallerBracket = 8
    elif numRows >= 4 and numRows < 8:
        smallerBracket = 4
    nonPullouts = smallerBracket - (numRows - smallerBracket)
    print(nonPullouts)
    # topbracket:
    # smallerbracket - (totalplayers - smallerbracket) = num of players stored in top bracket

    # numQuadrants = smallerbracket/4
    # count makes sure that players are being put into a different quadrant each time

    # for the rest of the players, they will be placed in pullout brackets
    # players of the higher flight will be matched with a player of the lower flight using a random number generator
    # need to check that randomly selected player is not a higher flight player and also not part of the same club
    # once all the players in the higher flight are matched with someone, the players of the lower flight can
    # be matched with other random players in the list as long as they are from different clubs

    # calculate num players from high flight
    # subtract
    numRemaining = len(players) - nonPullouts
    print(numRemaining)
    hf = 0
    mf = 0
    lf = 0
    remClub = {}
    remPlayers = []
    hflight = []
    mflight = []
    lflight = []
    for i in range(nonPullouts,len(players)):
        # count clubs for remaining players
        if players[i][2] in remClub:
            remClub[players[i][2]] += 1
        else:
            remClub[players[i][2]] = 1
        # count number of hf, mf, lf left
        if players[i][3] == flights[0]:
            hf += 1
            hflight.append(players[i])
        if players[i][3] == flights[1]:
            mf += 1
            mflight.append(players[i])
        if letter == 'B' or letter == 'C':
            if players[i][3] == flights[2]:
                lf += 1
                lfight.append(players[i])
    print(hflight)
    # print(mflight)
    for player in mflight:
        player.append(remClub[player[2]])
    # print("mflight with club count:",mflight)
    mflight = sorted(mflight, key=itemgetter(4), reverse=True)
    print("*"*30)
    print(mflight)
    remPlayers = hflight + mflight + lflight
    print("remaning players: ",remPlayers)

    # sum number of clubs greater than 5
    # will be used later to offset our randint
    sumClubs5 = 0
    for club, numClubs in remClub.items():
        if numClubs > 5:
            sumClubs5 += numClubs
    print("*"*30)
    print(sumClubs5)
    # print(remPlayers)

    print(hf, mf, lf)
    pulloutmatches = []
    haveopponents = [None]*len(remPlayers)
    matches = 0
    print(players)
    if lf == 0:
        if hf<mf:
            # start from front
            for i in range(len(remPlayers)):
                if matches==16:
                    break
                if haveopponents[i]==i:
                    continue
                opponent = random.randint(hf+sumClubs5,len(remPlayers)-1)
                while haveopponents[opponent]==opponent or remPlayers[i][2]==remPlayers[opponent][2]:
                    opponent = random.randint(hf+sumClubs5,len(remPlayers)-1)

                haveopponents[opponent]=opponent
                haveopponents[i]=i
                print(haveopponents)
                pulloutmatches.append([remPlayers[i]]+[remPlayers[opponent]])
                matches+=1


    for match in pulloutmatches:
        print(match)
            # start from back

    count = 0
    topbracket = 0 #players who do not have a bye
    oddIndex = [0,7,4,3,5,2,6,1]
    evenIndex = [7,0,3,4,2,5,1,6]
    i = 0
    pullouts = []
    players = players[0:nonPullouts]
    players = players + pulloutmatches
    print(len(players))
    print(players)
    for j in range(len(players)):
        print(i)
        # if topbracket >= nonPullouts:
        #     break
        if count == 0:
            thirdquadrant[oddIndex[i]] = players[j]
            count = 1
        elif count == 1:
            secondquadrant[evenIndex[i]] = players[j]
            count = 2
        elif count == 2:
            fourthquadrant[evenIndex[i]] = players[j]
            count = 3
        elif count == 3:
            firstquadrant[oddIndex[i]] = players[j]
            count = 0
            i += 1
        topbracket += 1
        print(players[j])
    print("topbracket: ", topbracket)

    # for i in range(topbracket,len(players)):

    # print("*****************")
    # print(firstquadrant)
    # print("*****************")
    # print(secondquadrant)
    # print("*****************")
    # print(thirdquadrant)
    # print("*****************")
    # print(fourthquadrant)

    draw = firstquadrant + secondquadrant + thirdquadrant + fourthquadrant
    print(draw)

    printDraw(players, draw)

    #notes for tomorrow: change it so it the low flight players are playing high flight players. Suggestions: make the quadrants have a fixed length of 8 then put in players from strong to weak as far as possible. so first two entered will be on opposite sides then next will go middle then next will go in between first and middle and next will go in between middle and last etc until it is full.

def printDraw(players, draw):
    drawTemplate = load_workbook("drawTemplate.xltx")
    sheets = drawTemplate.sheetnames
    if len(players) <= 64:
        source = drawTemplate[sheets[1]]
        sheet = drawTemplate.copy_worksheet(source)
    print(sheet.title)

    curRow = 8
    print(sheet[3][0].value)
    for player in draw:
        print(player)
        # inner match
        if len(player) != 2:
            sheet[curRow][2].value = player[3] + " " + player[2] + " " + player[0] + " " + player[1]
        # pullout match
        else:
            sheet[curRow-3][0].value = player[0][3] + " " + player[0][2] + " " + player[0][0] + " " + player[0][1]
            sheet[curRow+2][0].value = player[1][3] + " " + player[1][2] + " " + player[1][0] + " " + player[1][1]
        curRow += 9

    drawTemplate.save("drawTemplateFinal.xltx")

def main():
    makeDraw()


if __name__ == '__main__':
    main()
