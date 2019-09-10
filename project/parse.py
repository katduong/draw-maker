from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font


def parseFile():
    wb = load_workbook('players.xlsx')
    main = wb['Players']

    flights = {
        'AMD': 0, 'BMD': 0,'CMD': 0,'DMD': 0,'AWD': 0,
        'BWD': 0,'CWD': 0,'DWD': 0,'AMX': 0,'BMX': 0,
        'CMX': 0,'DMX': 0,'AMS': 0,'BMS': 0,'CMS': 0,
        'DMS': 0,'AWS': 0,'BWS': 0,'CWS': 0,'DWS': 0,
        'AXD': 0,'BXD': 0,'CXD': 0,'DXD': 0
    }

    header = list(main.rows)[3]
    for cell in header:
        cell.font = Font(bold=True)
    # loop through every row and create a new sheet for each event
    # player will be appended to the event's sheet
    for row in main.rows:
        print(row[4].value)
        events = row[4].value
        # skip first couple lines
        if (events != None):
            eventList = events.split(", ")
            print(eventList)
            for event in eventList:
                if event in flights:
                    if flights[event] == 0:
                        flights[event] += 1
                        newSheet = wb.create_sheet(event)
                        newSheet.append(cell.value for cell in header)
                        newSheet.append(cell.value for cell in row)
                    else:
                        wb[event].append(cell.value for cell in row)
    # sort sheets in alphabetical order
    wb._sheets.sort(key=lambda ws:ws.title)
    wb.save('playerParse.xlsx')

def reformatParse():
    wb = load_workbook("playerParse.xlsx")
    for sheet in wb:
        if sheet.title != "Players":
            sheet.insert_cols(5, amount=1)
            # if sheet is singles, don't need to worry about partners
            # just need to add reformat columns
            # lastname, firstname, gender, event letter(s), club
            for row in sheet:
                playerEvents = {
                    'MS': "",
                    'WS': "",
                    'MX': "",
                    'MD': "",
                    'WD': "",
                    'XD': ""
                }
                events = row[5].value.split(", ")
                # exclude first row
                if events != ['Events']:
                    for event in events:
                        # store event letters in player events dictionary
                        string = event[-2:]
                        playerEvents[string] += event[0]
                    row[4].value = playerEvents[sheet.title[1:]]
                else:
                    row[4].value = "Events"
                print(sheet.title)
            # doubles event
            # need to create a column for doubles partners
            # lastname, firstname, gender, event letter(s), doubles partner, club
            if 'MD' in sheet.title or 'WD' in sheet.title or 'MX' in sheet.title or 'XD' in sheet.title:
                # insert a column for the partner name
                sheet.insert_cols(6, amount=1)
                # insert a column for the partner club
                sheet.insert_cols(7, amount=1)
                players = {}
                rowCount = 1
                for row in sheet:
                    if rowCount == 1:
                        rowCount += 1
                        continue
                    # grab entry info for that player
                    entryInfo = row[8].value
                    currentPlayer = row[1].value + " " + row[0].value
                    print("*"*20)
                    print(currentPlayer)
                    print("entry info:", entryInfo)
                    if entryInfo != None:
                        # gets entry info and splits into list per each event
                        entryInfo = entryInfo.split("\n")
                        # remove empty string from the last index of list
                        if entryInfo[-1] == "":
                            entryInfo.pop()
                        print("entry info after pop()",entryInfo)
                        # loop through each event in the player's entry entryInfo
                        # store their partner's name in player dictionary
                        for event in entryInfo:
                            if sheet.title in event:
                                print("event:",event)
                                # partnerName contains first and last name
                                partnerName = event[4:].split(" (")[0]
                                partnerName = partnerName.split(" ")
                                print("partner name:",partnerName)
                                # stores each player as a key in the dictionary
                                # values will be a list of lists
                                # [[partner, names], [club, rowIndex]]
                                players[currentPlayer] = []
                                players[currentPlayer].append([])
                                for name in partnerName:
                                    players[currentPlayer][0].append(name)
                                club = row[3].value
                                players[currentPlayer].append([])
                                players[currentPlayer][1].append(club)
                                players[currentPlayer][1].append(rowCount)
                print(players)
                # iterate through each row again and check if player is listed in the dictionary another player's partners
                # if a match is found, append to the first occurence of pair's dictionary, the partner's club name
                rowCount = 1
                for row in sheet:
                    if rowCount == 1:
                        rowCount += 1
                        continue
                    playerName = row[1].value + " " + row[0].value
                    if playerName not in players:
                        print(playerName, "not found in player dictionary")
                        continue
                    # if player's partner in dictionary doesn't have a first name or last name, skip
                    if players[playerName][0][0] == "":
                        continue
                    #  check if player's partner only has a first name listed
                    elif players[playerName][0][-1] == players[playerName][0][0]:
                        partnerfn = players[playerName][0][0]
                        partnerFullName = partnerfn
                    #  otherwise, we know their partner has a first and last name
                    else:
                        partnerfn = players[playerName][0][0]
                        partnerln = players[playerName][0][-1]
                        partnerFullName = partnerfn + " " + partnerln

                    # check if partner full name matches with a key in the dictionary
                    if partnerFullName in players:
                        playerFnInDict = players[partnerFullName][0][0]
                        playerLnInDict = players[partnerFullName][0][-1]
                        # if playerFnInDict == row[1].value:
                            # store partner name in excel sheet
                        row[5].value = partnerFullName
                        if players[partnerFullName][1][0] != None:
                            # store partner's club name in excel sheet
                            row[6].value = players[partnerFullName][1][0]
                        # elif playerLnInDict == row[0].value:
                        #     row[5].value = partnerFullName
                        #     if players[partnerFullName][1][0] != None:
                        #         # store partner's club name in excel sheet
                        #         row[6].value = players[partnerFullName][1][0]




                    # if partnerFullName:
                    #     print(partnerFullName)
                    # for key in players:
                    #     if key == partner
                    #     playerFnInDict = player[key][0][0]
                    #     playerLnInDict = player[key][0][-1]
                    #     if partnerfn == playerFnInDict:


                    rowCount += 1

                            #     partner = players.get(partnerName)
                            #     if partner != None:
                            #         partner = partner[0].split(" ")[0]
                            # else:
                            #     continue
                            # if partner != None:
                            #     print("partner in dictionary",partner)
                            # # if row[1].value in players.values():
                            # if partner != None and partner == row[1].value:
                            #     print("row[1].value: ", row[1].value, " ==", partner )
                            #
                            #     # insert partner's club name into dictionary
                            #     print("partner's clubname: ", row[3].value, partner)
                            #     players[partnerName].append(row[3].value)
                            #     sheet[players[partnerName]][5].value = players[partnerName][1]
                            #     # row[5].value = players[partnerName][1]
                            #     if players[partnerName][2]:
                            #         # row[6].value = players[partnerName][2]
                            #         sheet[players[partnerName][0]][6].value = players[partnerName][2]
                            #     print("player key:",players[partnerName][1])
                            #     # sheet.delete_rows(rowCount,1)
                            # else:
                            #     # store the first occurence of a pair in the dictionary
                            #     players[row[1].value + " " + row[0].value] = []
                            #     # players[row[1].value + " "  + row[0].value].append(rowCount)
                            #     players[row[1].value + " " + row[0].value].append(partnerName)
                            #
                    rowCount += 1
                    # index = entryInfo.find(sheet.title)
                    # partnerIndex = index += 4
                    # print(partnerIndex)
                    # partnerName = entryInfo[partnerIndex:]


    wb.save("playerList.xlsx")


def main():
    # parseFile()
    reformatParse()


if __name__ == '__main__':
    main()
